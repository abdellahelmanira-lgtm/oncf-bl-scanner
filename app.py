import streamlit as st
import json
import base64
import requests
from datetime import datetime
from PIL import Image, ImageEnhance
import io
import openpyxl
import os
import re
import pandas as pd

# --- إعدادات الصفحة ---
st.set_page_config(page_title="ONCF | Scanner BL", page_icon="🚆", layout="centered")

# --- 🔑 بلاصة API KEY ديالك ---
API_KEY = "AIzaSyBfmryF5dRnwFLtMSJJknFPF3d4Ky_SDzo" 
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent?key={API_KEY}"

# --- حفظ البيانات فالذاكرة ديال التطبيق (Session State) ---
if "extracted_data" not in st.session_state:
    st.session_state.extracted_data = None
if "excel_file_data" not in st.session_state:
    st.session_state.excel_file_data = None

# --- البرومبت المترجم بالحرف ---
EXTRACTION_PROMPT = """
You are an expert OCR and document analysis system specialized in ONCF (Office National des Chemins de Fer du Maroc) delivery documents called "Bordereau de Livraison" (BL).
Your task: Carefully read the provided document image and extract ALL visible information into a strictly structured JSON object.
CRITICAL RULES:
1. Return ONLY a valid JSON object — no markdown, no explanation, no code blocks.
2. If a field is not visible or not found, use null.
3. For the items table, extract EVERY row visible — do not skip any.
4. Preserve original French text exactly as written.
5. For dates, preserve the original format found in the document.
6. Numbers should be returned as numbers (not strings) where they represent quantities.
7. CRITICAL: For 'sender.name' and 'recipient.name', NEVER use the full phrase "OFFICE NATIONAL DES CHEMINS DE FER". ALWAYS extract ONLY the specific short abbreviation/code (e.g., EMC, EMIC, EMS, TMK, CLC, TMFRET) found in the header, text, or stamps.
8. STRICT SPLIT RULE: Check the "Quantité" column FIRST. If the quantity is 1 (or "01"), you MUST NEVER split the item into multiple rows, no matter how many lines the observation text takes. Only split if Quantité > 1 AND you see multiple distinct serial numbers.
9. SERIAL NUMBER (CODE) EXTRACTION: The text on these documents often uses a printed script font that resembles cursive handwriting. Do not ignore it. Look carefully in the "DETAIL DES MARCHANDISES" column, directly next to or below the main designation (e.g., next to "Pantographe"). 
10. When you find a value like "N°FS52/92" written in this script font, extract it into the "code" field, but remove the "N°" prefix (e.g., output "FS52/92" or "2459/18").
11. MULTI-LINE OBSERVATIONS: If an item has a quantity of 1, combine all lines of text in the "OBSERVATIONS" column (e.g., "Bras cassé oxydation coincement mecanique Z2M116") into a single string for that one item.

{
  "documentNumber": "string or null", "date": "string or null", "reference": "string or null",
  "sender": {"name": "string or null", "department": "string or null"},
  "recipient": {"name": "string or null", "department": "string or null"},
  "observations": "string or null",
  "items": [{"lineNumber":1,"code":null,"designation":null,"quantity":null,"unit":null,"status":null,"observations":null}],
  "confidenceNote": "string"
}
Extract now:
"""

def process_image(image):
    img = image.convert('L')
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = ImageEnhance.Sharpness(img).enhance(1.5)
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='JPEG', quality=90)
    return base64.b64encode(img_byte_arr.getvalue()).decode("utf-8")

def extract_data(base64_image):
    payload = {
        "contents": [{"parts": [{"text": EXTRACTION_PROMPT}, {"inline_data": {"mime_type": "image/jpeg", "data": base64_image}}]}],
        "generationConfig": {"temperature": 0.1, "topP": 0.9}
    }
    response = requests.post(GEMINI_URL, json=payload, headers={"Content-Type": "application/json"})
    
    if response.status_code == 200:
        raw_text = response.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
        raw_text = raw_text.replace("```json", "").replace("```", "").strip()
        try:
            return json.loads(raw_text)
        except json.JSONDecodeError:
            match = re.search(r'\{[\s\S]*\}', raw_text)
            if match:
                try:
                    return json.loads(match.group(0))
                except:
                    pass
            st.error("Gemini n'a pas retourné un JSON valide.")
            return None
    else:
        st.error(f"Erreur API: {response.text}")
        return None

def update_master_excel(data, template_path="BLs Etablissements ONCF 01-01-2026.xlsx"):
    if not os.path.exists(template_path):
        st.error(f"⚠️ الملف الأصلي '{template_path}' ما كاينش! تأكد أنك حطيتيه فـ GitHub.")
        return None

    wb = openpyxl.load_workbook(template_path)

    def get_nom(champ):
        if not champ: return ""
        if isinstance(champ, str): return champ
        if isinstance(champ, dict):
            return champ.get('nom') or champ.get('name') or champ.get('entreprise') or champ.get('code') or (list(champ.values())[0] if champ else "")
        return str(champ)

    exp = get_nom(data.get('sender'))
    dest = get_nom(data.get('recipient'))

    expediteur = str(exp).strip().upper() if exp else "INCONNU"
    destinataire = str(dest).strip().upper() if dest else "INCONNU"

    sheet_name1 = f"{expediteur}->{destinataire}"
    sheet_name2 = f"{expediteur}-{destinataire}"
    
    sheet = None
    if sheet_name1 in wb.sheetnames: sheet = wb[sheet_name1]
    elif sheet_name2 in wb.sheetnames: sheet = wb[sheet_name2]

    if not sheet:
        st.error(f"⚠️ مالقيناش الورقة ديال '{sheet_name1}' ولا '{sheet_name2}' فالملف الأصلي! تأكد من المُرسل والمُستقبل.")
        return None

    articles = data.get('items') or []
    date_bl = data.get('date') or ""
    num_bl = data.get('reference') or data.get('documentNumber') or ""

    real_last_row = 6
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value or sheet.cell(row=i, column=2).value:
            real_last_row = i

    for item in articles:
        real_last_row += 1
        final_date = item.get('date') or date_bl
        final_num_bl = item.get('reference') or num_bl or item.get('documentNumber') or ""

        sheet.cell(row=real_last_row, column=1, value=final_date)
        sheet.cell(row=real_last_row, column=2, value=item.get('designation', ""))
        sheet.cell(row=real_last_row, column=3, value=item.get('code') or "")
        qty = item.get('quantity')
        sheet.cell(row=real_last_row, column=4, value=qty if qty is not None and str(qty).strip() != "" else 1)
        sheet.cell(row=real_last_row, column=5, value=final_num_bl)
        sheet.cell(row=real_last_row, column=6, value=item.get('observations') or "")

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()

# --- واجهة المستخدم ---
st.markdown("<h1 style='text-align: center; color: #002E5D;'>🚆 ONCF | OCR & Injection</h1>", unsafe_allow_html=True)
st.divider()

col1, col2 = st.columns(2)
with col1: upload_img = st.file_uploader("📁 Galerie", type=["jpg", "jpeg", "png"])
with col2: camera_img = st.camera_input("📸 Caméra")

image_to_process = camera_img if camera_img else upload_img

if image_to_process:
    st.image(image_to_process, caption="Image à traiter", use_column_width=True)
    
    if st.button("⚡ 1. Extraire les données", use_container_width=True, type="primary"):
        with st.spinner("⏳ Extraction par l'IA en cours..."):
            img = Image.open(image_to_process)
            b64_img = process_image(img)
            data = extract_data(b64_img)
            
            if data:
                st.session_state.extracted_data = data
                st.session_state.excel_file_data = None # مسح أي ملف قديم
                st.success("✅ Extraction réussie ! Vérifiez les données ci-dessous.")

# --- واجهة المراجعة والتعديل ---
if st.session_state.extracted_data:
    st.markdown("### 📝 Étape 2: Vérification et Modification")
    data = st.session_state.extracted_data
    
    def get_dict_val(d, key):
        if isinstance(d, dict): return d.get(key, "")
        return str(d) if d else ""

    # تعديل المعلومات العامة
    col_a, col_b = st.columns(2)
    with col_a:
        new_exp = st.text_input("Expéditeur (ex: EMC)", value=get_dict_val(data.get("sender"), "name"))
        new_date = st.text_input("Date du BL", value=data.get("date", ""))
    with col_b:
        new_dest = st.text_input("Destinataire (ex: EMIC)", value=get_dict_val(data.get("recipient"), "name"))
        new_ref = st.text_input("N° BL (Reference)", value=data.get("reference") or data.get("documentNumber") or "")

    # تعديل جدول السلعة
    st.markdown("#### 📦 Articles (Modifiez directement dans le tableau)")
    
    # تحويل السلعة لـ DataFrame باش تخدم فـ data_editor
    items = data.get("items", [])
    df_items = pd.DataFrame(items)
    
    # التأكد أن الأعمدة المطلوبة كاينا واخا تكون خاوية
    for col in ["designation", "code", "quantity", "observations"]:
        if col not in df_items.columns:
            df_items[col] = ""
            
    # عرض الجدول القابل للتعديل
    edited_df = st.data_editor(df_items, num_rows="dynamic", use_container_width=True)

    # زر الحفظ والحقن
    if st.button("💾 3. Valider et Injecter dans Excel", use_container_width=True):
        with st.spinner("⏳ Injection en cours..."):
            # تحديث البيانات بالمعلومات الجديدة اللي دخلتي
            updated_data = {
                "sender": {"name": new_exp},
                "recipient": {"name": new_dest},
                "date": new_date,
                "reference": new_ref,
                "items": edited_df.to_dict(orient="records")
            }
            
            # حقن فـ Excel
            excel_data = update_master_excel(updated_data)
            
            if excel_data:
                st.session_state.excel_file_data = excel_data
                st.success("✅ Fichier Excel mis à jour avec succès !")

# --- زر التحميل ---
if st.session_state.excel_file_data:
    st.download_button(
        label="📥 Télécharger le fichier de travail",
        data=st.session_state.excel_file_data,
        file_name=f"BL_ONCF_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
